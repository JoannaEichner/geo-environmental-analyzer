from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from geo_environmental_analyzer.application.run_pipeline import RunAnalysisPipeline
from geo_environmental_analyzer.domain.models import (
    GroundWaterResult,
    OrderedRoute,
    ParcelRecord,
    ProtectedAreaDistance,
    SurfaceWaterResult,
)
from geo_environmental_analyzer.infrastructure.input.points_txt import TxtPointReader
from geo_environmental_analyzer.infrastructure.reporting.xlsx_writer import (
    XlsxReportWriter,
)


class StubParcelAnalyzer:
    def analyze(self, route: OrderedRoute) -> list[ParcelRecord]:
        assert len(route.points) == 2
        return [
            ParcelRecord(
                parcel_number="194",
                cadastral_district_code="0015",
                cadastral_district_name="Jezowka",
                municipality_name="Sochaczew",
                parcel_identifier="142807_2.0015.194",
            )
        ]


class StubSurfaceWaterAnalyzer:
    def analyze(self, route: OrderedRoute) -> list[SurfaceWaterResult]:
        assert len(route.points) == 2
        return [
            SurfaceWaterResult(
                code="PLRW2000112727699",
                name="Pisia test",
                status="naturalna czesc wod",
                monitored="TAK",
                overall_state="zly stan wod",
                risk_assessment="zagrozona",
                ecological_potential_goal="umiarkowany stan ekologiczny",
                chemical_goal="dobry stan chemiczny",
            )
        ]


class StubGroundWaterAnalyzer:
    def analyze(self, route: OrderedRoute) -> list[GroundWaterResult]:
        assert len(route.points) == 2
        return [
            GroundWaterResult(
                code="PLGW200065",
                name="Jednolita czesc wod podziemnych nr 65 PLGW200065",
                monitored="TAK",
                chemical_state="dobry",
                quantitative_state="dobry",
                overall_state="dobry",
                risk_assessment="niezagrozona",
                quantitative_goal="dobry stan ilosciowy",
                chemical_goal="dobry stan chemiczny",
            )
        ]


class StubProtectedAreaAnalyzer:
    def analyze(self, route: OrderedRoute) -> list[ProtectedAreaDistance]:
        assert len(route.points) == 2
        return [
            ProtectedAreaDistance(
                form_name='Rezerwat przyrody "Wielkie Jezioro"',
                distance_km=1.23,
            )
        ]


def test_pipeline_and_writer_generate_expected_xlsx(tmp_path: Path) -> None:
    input_path = tmp_path / "points.txt"
    output_path = tmp_path / "report.xlsx"
    input_path.write_text(
        "1\tP1\t7450669.31\t5780467.22\n2\tP2\t7450700.00\t5780500.00\n",
        encoding="utf-8",
    )

    pipeline = RunAnalysisPipeline(
        point_reader=TxtPointReader(),
        parcel_analyzer=StubParcelAnalyzer(),
        surface_water_analyzer=StubSurfaceWaterAnalyzer(),
        ground_water_analyzer=StubGroundWaterAnalyzer(),
        protected_area_analyzer=StubProtectedAreaAnalyzer(),
    )

    bundle = pipeline.run(input_path)

    assert len(bundle.route.points) == 2
    assert len(bundle.parcels) == 1
    assert len(bundle.surface_water) == 1
    assert len(bundle.groundwater) == 1
    assert len(bundle.protected_areas) == 1

    XlsxReportWriter().write(bundle, output_path)

    assert output_path.exists()

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "01_Dzialki",
        "02_Wody_Status",
        "03_Wody_Cele",
        "04_Ochrona",
    ]

    parcels_sheet = workbook["01_Dzialki"]
    assert parcels_sheet["A1"].value == "Numer działki"
    assert parcels_sheet["B1"].value == "Obręb"
    assert parcels_sheet["A2"].value == "194"
    assert parcels_sheet["B2"].value == "0015 Jezowka"
    assert parcels_sheet["A2"].alignment.horizontal == "center"
    assert parcels_sheet["A2"].border.left.style == "thin"

    water_status_sheet = workbook["02_Wody_Status"]
    assert water_status_sheet["A1"].value == 'PLRW2000112727699 "Pisia test":'
    assert water_status_sheet["A2"].value == "Status JCWP"
    assert water_status_sheet["B2"].value == "naturalna czesc wod"
    assert (
        water_status_sheet["A7"].value
        == "Jednolita czesc wod podziemnych nr 65 PLGW200065"
    )
    assert water_status_sheet["B2"].alignment.horizontal == "center"
    assert water_status_sheet["B2"].border.right.style == "thin"

    water_goals_sheet = workbook["03_Wody_Cele"]
    assert water_goals_sheet["A1"].value == 'PLRW2000112727699 "Pisia test":'
    assert water_goals_sheet["A2"].value == "stan/potencjał ekologiczny"
    assert water_goals_sheet["B2"].value == "umiarkowany stan ekologiczny"

    protection_sheet = workbook["04_Ochrona"]
    assert protection_sheet["A1"].value == "Nazwa formy ochrony przyrody"
    assert protection_sheet["B1"].value == "Odległość od planowanej inwestycji"
    assert protection_sheet["A2"].value == 'Rezerwat przyrody "Wielkie Jezioro"'
    assert protection_sheet["B2"].value == "1,23 km"
    assert protection_sheet["B2"].alignment.horizontal == "center"
    assert protection_sheet["B2"].border.bottom.style == "thin"
