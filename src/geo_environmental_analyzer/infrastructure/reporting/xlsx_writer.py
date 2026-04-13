from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from geo_environmental_analyzer.domain.models import AnalysisBundle, ParcelRecord
from geo_environmental_analyzer.domain.protocols import ReportWriter

THIN_SIDE = Side(style="thin", color="000000")
TABLE_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


@dataclass(slots=True)
class ParcelSheetRow:
    parcel_number: str
    cadastral_district: str


def map_parcels_to_sheet_rows(parcels: list[ParcelRecord]) -> list[ParcelSheetRow]:
    rows: list[ParcelSheetRow] = []

    for parcel in parcels:
        district_parts = [
            part
            for part in [parcel.cadastral_district_code, parcel.cadastral_district_name]
            if part
        ]
        rows.append(
            ParcelSheetRow(
                parcel_number=parcel.parcel_number,
                cadastral_district=" ".join(district_parts),
            )
        )

    return rows


class XlsxReportWriter(ReportWriter):
    def write(self, bundle: AnalysisBundle, output_path: Path) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        self._write_parcels_sheet(workbook, bundle)
        self._write_water_status_sheet(workbook, bundle)
        self._write_water_goals_sheet(workbook, bundle)
        self._write_protected_areas_sheet(workbook, bundle)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def _write_parcels_sheet(self, workbook: Workbook, bundle: AnalysisBundle) -> None:
        sheet = workbook.create_sheet("01_Dzialki")
        sheet.append(["Numer działki", "Obręb"])

        for row in map_parcels_to_sheet_rows(bundle.parcels):
            sheet.append([row.parcel_number, row.cadastral_district])

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 38

        self._style_header_row(sheet, 1)
        self._style_table(sheet, 1, sheet.max_row, 1, 2)

    def _write_water_status_sheet(
        self, workbook: Workbook, bundle: AnalysisBundle
    ) -> None:
        sheet = workbook.create_sheet("02_Wody_Status")
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 72

        for item in bundle.surface_water:
            self._append_block(
                sheet,
                f'{item.code} "{item.name}":',
                [
                    ("Status JCWP", item.status),
                    ("monitorowana", self._normalize_monitoring(item.monitored)),
                    ("stan (ogólny)", item.overall_state),
                    (
                        "ocena ryzyka nieosiągnięcia celów środowiskowych",
                        self._normalize_risk(item.risk_assessment),
                    ),
                ],
            )

        for item in bundle.groundwater:
            self._append_block(
                sheet,
                item.name,
                [
                    ("monitorowana", self._normalize_monitoring(item.monitored)),
                    ("stan chemiczny", item.chemical_state),
                    ("stan ilościowy", item.quantitative_state),
                    ("stan JCWPd", item.overall_state),
                    (
                        "ocena ryzyka nieosiągnięcia celów środowiskowych",
                        self._normalize_risk(item.risk_assessment),
                    ),
                ],
            )

    def _write_water_goals_sheet(
        self, workbook: Workbook, bundle: AnalysisBundle
    ) -> None:
        sheet = workbook.create_sheet("03_Wody_Cele")
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 92

        for item in bundle.surface_water:
            self._append_block(
                sheet,
                f'{item.code} "{item.name}":',
                [
                    ("stan/potencjał ekologiczny", item.ecological_potential_goal),
                    ("stan chemiczny", item.chemical_goal),
                ],
            )

        for item in bundle.groundwater:
            self._append_block(
                sheet,
                item.name,
                [
                    ("stan chemiczny", item.chemical_goal),
                    ("stan ilościowy", item.quantitative_goal),
                ],
            )

    def _write_protected_areas_sheet(
        self,
        workbook: Workbook,
        bundle: AnalysisBundle,
    ) -> None:
        sheet = workbook.create_sheet("04_Ochrona")
        sheet.append(
            ["Nazwa formy ochrony przyrody", "Odległość od planowanej inwestycji"]
        )

        for item in bundle.protected_areas:
            sheet.append([item.form_name, self._format_distance(item.distance_km)])

        sheet.column_dimensions["A"].width = 72
        sheet.column_dimensions["B"].width = 26

        self._style_header_row(sheet, 1)
        self._style_table(sheet, 1, sheet.max_row, 1, 2)

    def _append_block(
        self,
        sheet: Worksheet,
        title: str,
        rows: list[tuple[str, str]],
    ) -> None:
        is_empty_sheet = (
            sheet.max_row == 1
            and sheet["A1"].value is None
            and sheet["B1"].value is None
        )
        start_row = 1 if is_empty_sheet else sheet.max_row + 2
        end_row = start_row + len(rows)

        sheet.cell(row=start_row, column=1, value=title)
        sheet.cell(row=start_row, column=2, value="")

        for row_index, (label, value) in enumerate(rows, start=start_row + 1):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)

        sheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=2,
        )

        self._style_block(sheet, start_row, end_row)

    def _style_block(self, sheet: Worksheet, start_row: int, end_row: int) -> None:
        title_cell = sheet.cell(row=start_row, column=1)
        title_cell.font = Font(bold=True)
        title_cell.fill = TITLE_FILL
        title_cell.alignment = CENTER_ALIGNMENT
        title_cell.border = TABLE_BORDER

        trailing_title_cell = sheet.cell(row=start_row, column=2)
        trailing_title_cell.fill = TITLE_FILL
        trailing_title_cell.border = TABLE_BORDER

        for row in range(start_row + 1, end_row + 1):
            for column in range(1, 3):
                cell = sheet.cell(row=row, column=column)
                cell.alignment = CENTER_ALIGNMENT
                cell.border = TABLE_BORDER

        for row in range(start_row, end_row + 1):
            sheet.row_dimensions[row].height = 24

    def _style_header_row(self, sheet: Worksheet, row_index: int) -> None:
        for cell in sheet[row_index]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGNMENT
            cell.border = TABLE_BORDER
        sheet.row_dimensions[row_index].height = 26

    def _style_table(
        self,
        sheet: Worksheet,
        start_row: int,
        end_row: int,
        start_column: int,
        end_column: int,
    ) -> None:
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                cell = sheet.cell(row=row, column=column)
                cell.alignment = CENTER_ALIGNMENT
                cell.border = TABLE_BORDER

        for row in range(start_row + 1, end_row + 1):
            sheet.row_dimensions[row].height = 22

    def _format_distance(self, value: float) -> str:
        return f"{value:.2f}".replace(".", ",") + " km"

    def _normalize_monitoring(self, value: str) -> str:
        normalized = value.strip().lower()
        if normalized in {"tak", "nie"}:
            return normalized
        return value

    def _normalize_risk(self, value: str) -> str:
        normalized = value.strip().lower()
        if "niezagro" in normalized:
            return "niezagrożona"
        if "zagro" in normalized:
            return "zagrożona"
        return value
